from celery import shared_task
from openpyxl import load_workbook
from django.db import transaction
from .models import *

@shared_task(bind=True)
def process_import(self, job_id):
    
    # get ImportJob
    job = ImportJob.objects.get(id=job_id)
    job.status = 'processing'
    job.save()
    
    print('File path:', job.file.path)

    try:
        # open excel file
        wb = load_workbook(job.file.path)
        
        # getting current active sheets from the excel file
        ws = wb.active
        
        # read headers from the 1st row of the excel file
        headers = [cell.value for cell in ws[1]]
        print('Headers:', headers)
        
        # defining required header fields (static defined)
        # for now, all header fields are mandatory
        required_fields = ['sku', 'title', 'price', 'stock', 'description', 'image_url', 'brand']

        # Display the user, required missing field names
        for field in required_fields:
            if field not in headers:
                print('Missing header: ', field)
                job.status = 'failed'
                job.save()
                return
            
        # prepare product cache
        # comment: the efficient way could be, we can create index on "sku" field
        # we can find out / filter product based on sku field
        existing_products = {
            p.sku: p for p in Product.objects.only('sku')
        }
        
        # create list
        new_products = []
        update_products = []
        errors = []
        
        # counter
        total = 0
        success = 0
        failed = 0
        
        # loop excel rows
        for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            total += 1
            
            try:
                
                # convert row into dictionary
                data = dict(zip(headers, row))
                
                sku = str(data['sku']).strip()
                title = str(data['title']).strip()
                price = float(data['price'])
                stock = int(data['stock'])
                
                # decide create / update
                if sku in existing_products:
                    
                    product = existing_products[sku]
                    product.title = title
                    product.price = price
                    product.stock = stock
                    product.description = data.get('description')
                    product.image_url = data.get('image_url')
                    product.brand = data.get('brand')
                    
                    update_products.append(product)
                    
                else:
                    new_products.append(
                        Product(
                            sku=sku,
                            title=title,
                            price=price,
                            stock=stock,
                            description=data.get('description'),
                            image_url=data.get('image_url'),
                            brand=data.get('brand'),
                        )
                    )
                success += 1
            
            except Exception as e:
                failed += 1
                
                errors.append(
                    ImportError(
                        job=job,
                        row_number=row_number,
                        message=str(e),
                    )
                )
        
        # save everything in db (bulk)
        with transaction.atomic():
            
            if new_products:
                Product.objects.bulk_create(new_products)
                
            if update_products:
                Product.objects.bulk_update(
                    update_products,
                    ['title', 'price', 'stock', 'description', 'image_url', 'brand']
                )
            if errors:
                ImportError.objects.bulk_create(errors)
                
        # update job summary
        job.total_rows = total
        job.success_rows = success
        job.failed_rows = failed
        job.status = 'completed'
        job.save()
        
    except Exception as e:
        print('Task failed:', str(e))
        job.status = 'failed'
        job.save()